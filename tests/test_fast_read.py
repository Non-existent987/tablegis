#!/usr/bin/env python
# coding: utf-8
"""Tests for fast_read module"""

import os
import tempfile
import shutil
import pytest
import polars as pl
import pandas as pd

import tablegis as tg


@pytest.fixture
def sample_excel(tmp_path):
    """创建一个多sheet的测试Excel文件"""
    file_path = tmp_path / "test_data.xlsx"
    df1 = pl.DataFrame({
        "城市": ["北京", "上海", "广州", "深圳", "杭州"],
        "经度": [116.4, 121.5, 113.3, 114.1, 120.2],
        "纬度": [39.9, 31.2, 23.1, 22.5, 30.3],
        "需求ID": [1, 2, 3, 4, 5],
    })
    df2 = pl.DataFrame({
        "需求ID": [1, 2, 3, 4, 5],
        "类型": ["5G", "4G", "5G", "4G", "5G"],
        "频段": ["3.5G", "1.8G", "2.6G", "900M", "3.5G"],
    })
    # 写入多sheet Excel
    df1.write_excel(str(file_path), worksheet="站点信息")
    # 用openpyxl追加第二个sheet
    import openpyxl
    wb = openpyxl.load_workbook(str(file_path))
    ws = wb.create_sheet("基站信息")
    ws.append(["需求ID", "类型", "频段"])
    for row in df2.iter_rows(named=True):
        ws.append([row["需求ID"], row["类型"], row["频段"]])
    wb.save(str(file_path))
    wb.close()
    return str(file_path)


@pytest.fixture
def sample_csv(tmp_path):
    """创建一个测试CSV文件"""
    file_path = tmp_path / "test_data.csv"
    df = pl.DataFrame({
        "城市": ["北京", "上海", "广州"],
        "经度": [116.4, 121.5, 113.3],
        "纬度": [39.9, 31.2, 23.1],
    })
    df.write_csv(str(file_path))
    return str(file_path)


class TestFastReadExcel:

    def test_read_single_sheet(self, sample_excel):
        """读取单个sheet"""
        df = tg.fast_read(sample_excel, sheet="站点信息")
        assert isinstance(df, pl.DataFrame)
        assert len(df) == 5
        assert "城市" in df.columns

    def test_read_single_sheet_to_pandas(self, sample_excel):
        """读取单个sheet并转为pandas"""
        df = tg.fast_read(sample_excel, sheet="站点信息", to_pandas=True)
        assert isinstance(df, pd.DataFrame)
        assert len(df) == 5
        assert "城市" in df.columns

    def test_read_all_sheets(self, sample_excel):
        """读取全部sheet"""
        data = tg.fast_read(sample_excel)
        assert isinstance(data, dict)
        assert "站点信息" in data
        assert "基站信息" in data
        assert len(data["站点信息"]) == 5

    def test_read_with_columns(self, sample_excel):
        """只读取指定列"""
        df = tg.fast_read(sample_excel, sheet="站点信息", columns=["城市", "经度"])
        assert df.columns == ["城市", "经度"]
        assert len(df) == 5

    def test_read_cached(self, sample_excel):
        """第二次读取走缓存"""
        # 第一次
        df1 = tg.fast_read(sample_excel, sheet="站点信息")
        # 第二次（应走缓存，不打印转换信息）
        df2 = tg.fast_read(sample_excel, sheet="站点信息")
        assert df1.equals(df2)

    def test_read_nonexistent_sheet(self, sample_excel):
        """读取不存在的sheet应报错"""
        with pytest.raises(FileNotFoundError):
            tg.fast_read(sample_excel, sheet="不存在的sheet")

    def test_read_nonexistent_file(self):
        """读取不存在的文件应报错"""
        with pytest.raises(FileNotFoundError):
            tg.fast_read("nonexistent.xlsx", sheet="test")

    def test_read_second_sheet_only(self, sample_excel):
        """只读第二个sheet，不会转换第一个"""
        df = tg.fast_read(sample_excel, sheet="基站信息")
        assert isinstance(df, pl.DataFrame)
        assert "类型" in df.columns
        assert len(df) == 5

    def test_type_casting(self, sample_excel):
        """数字列应自动转为数值类型"""
        df = tg.fast_read(sample_excel, sheet="站点信息")
        # 经度纬度应为Float64，城市应为String
        assert df["经度"].dtype in (pl.Float64, pl.Float32)
        assert df["城市"].dtype == pl.String

    def test_refresh(self, sample_excel):
        """refresh=True 强制重新转换"""
        df1 = tg.fast_read(sample_excel, sheet="站点信息")
        df2 = tg.fast_read(sample_excel, sheet="站点信息", refresh=True)
        assert df1.equals(df2)


class TestFastReadCSV:

    def test_read_csv(self, sample_csv):
        """读取CSV文件"""
        df = tg.fast_read(sample_csv)
        assert isinstance(df, pl.DataFrame)
        assert len(df) == 3
        assert "城市" in df.columns

    def test_read_csv_to_pandas(self, sample_csv):
        """读取CSV并转为pandas"""
        df = tg.fast_read(sample_csv, to_pandas=True)
        assert isinstance(df, pd.DataFrame)
        assert len(df) == 3

    def test_read_csv_with_columns(self, sample_csv):
        """只读CSV的指定列"""
        df = tg.fast_read(sample_csv, columns=["城市"])
        assert df.columns == ["城市"]


class TestFastReadImport:

    def test_import_via_tablegis(self):
        """通过 tablegis 导入 fast_read"""
        assert hasattr(tg, 'fast_read')
        assert callable(tg.fast_read)

    def test_fast_read_has_docstring(self):
        """fast_read 应有文档字符串"""
        assert tg.fast_read.__doc__ is not None
        assert "按需" in tg.fast_read.__doc__ or "缓存" in tg.fast_read.__doc__
